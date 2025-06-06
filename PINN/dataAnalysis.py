import pandas as pd
import numpy as np
import random
import os
import openpyxl
import re
from decimal import Decimal

# 读取每个表格的数据并存储在字典中
def get_data(data_sheet_name, 
             sheet_name_template = "avg_1_z", 
             usecols=("B:EG"), 
             z_size = 72, 
             y_size = 136):
    if sheet_name_template.count('z') >= 2:
        print(f"模板字符串{sheet_name_template}包含两个或多个'z'字符")
        return 0
    data = {}
    if os.path.exists(data_sheet_name.split(".")[0]+"pkl_data"):
        for z in range(z_size):
            if os.path.exists(data_sheet_name.split(".")[0]+"pkl_data/pkl%d"%(z)):
                df = pd.read_pickle(data_sheet_name.split(".")[0]+"pkl_data/pkl%d"%(z))
            else:
                sheet_name = sheet_name_template.replace("z", f"{z+1}")
                df = pd.read_excel(data_sheet_name, sheet_name=sheet_name, header=None, usecols=usecols, names=list(range(y_size)))
                pd.to_pickle(df, data_sheet_name.split(".")[0]+"pkl_data/pkl%d"%(z))
            data[z] = df
    else:
        os.mkdir(data_sheet_name.split(".")[0]+"pkl_data")
        for z in range(z_size):
            sheet_name = sheet_name_template.replace("z", f"{z+1}")
            df = pd.read_excel(data_sheet_name, sheet_name=sheet_name, header=None, usecols=usecols, names=list(range(y_size)))
            pd.to_pickle(df, data_sheet_name.split(".")[0]+"pkl_data/pkl%d"%(z))
            data[z] = df
    return data

# 读取fluka导出的bnn.lis文件并转换为对应格式excel文件
def extract_matrix_values(file_path, output_path, activity = 1e11):
    with open(file_path, 'r') as file:
        lines = file.read()  # 读取整个文件为一个字符串

    # 使用正则表达式查找矩阵尺寸信息
    size_pattern = r'X coordinate: from.*?(\d+) bins.*?\n.*?Y coordinate: from.*?(\d+) bins.*?\n.*?Z coordinate: from.*?(\d+) bins'
    match = re.search(size_pattern, lines, re.DOTALL)
    
    nx, ny, nz = map(int, match.groups())

    # 使用正则表达式提取值矩阵的纯数值行
    value_matrix_pattern = r'(?=accurate deposition along the tracks requested)(?:.*)(?=Percentage errors follow in a matrix)'
    match = re.search(value_matrix_pattern, lines, re.DOTALL)
    print(match)
    matrix_values = re.findall(r'\d+\.\d*E[+-]?\d+', match.group())
    print(matrix_values)
    
    # 将字符串转换为 Decimal 对象进行运算
    decimal_values = [Decimal(x) for x in matrix_values]

    # 将 activity 和 1.602176462E-7*3600 转换为 Decimal
    activity = Decimal(activity)
    factor = Decimal(1.602176462e-7*36000000)

    # 格式与单位转换（默认fluka输出剂量率单位是Gev/s/primary，我要转换为mGy/h，要乘以1.602176462E-7*3600000*每秒粒子数（假设为1e10））
    result = [
        f"{x * activity * factor:.8f}"  # 保留 8 位有效数字
        for x in decimal_values
    ]
    
    # 导出为excel
    workbook = openpyxl.Workbook()

    for iz in range(nz):
        worksheet = workbook.create_sheet(title=f'idx={iz+1}')

        for iy in range(ny):
            row_values = [iy]+result[iz * ny * nx + iy * nx: iz * ny * nx + (iy + 1) * nx]
            for ix, value in enumerate(row_values, start=1):
                worksheet.cell(row=iy + 1, column=ix, value=float(value))

    workbook.save(output_path)

def training_sampling(data, 
                      center_x, 
                      center_y, 
                      center_z, 
                      inner_radius_max, 
                      vertice = False, 
                      inner_radius_list = list(range(1,3)), 
                      use_box_area = False, #是否使用固定空间来进行取样
                      use_box = [[0,0,0],99,74,99], #左前下点，x方向延申长度，y方向延申长度，z方向延申长度
                      step_sizes = [2,3,4], 
                      x_y_reverse = True, 
                      direction = "6vector",
                      sourcepos = [[48, 45, 5], [97, 90, 54]],#我们把源附近半径为10的球割掉
                      ):
    
    sampled_data = []
    
    if vertice:
        # 取样八个顶点的数据
        vertices = [item for sublist in
            [[(center_x + inner_radius, center_y + inner_radius, center_z + inner_radius),
            (center_x + inner_radius, center_y + inner_radius, center_z - inner_radius),
            (center_x + inner_radius, center_y - inner_radius, center_z + inner_radius),
            (center_x + inner_radius, center_y - inner_radius, center_z - inner_radius),
            (center_x - inner_radius, center_y + inner_radius, center_z + inner_radius),
            (center_x - inner_radius, center_y + inner_radius, center_z - inner_radius),
            (center_x - inner_radius, center_y - inner_radius, center_z + inner_radius),
            (center_x - inner_radius, center_y - inner_radius, center_z - inner_radius)]
            for inner_radius in inner_radius_list] for item in sublist
        ]
        for vertex in vertices:
            x, y, z = vertex
            if x_y_reverse == False:
                sampled_value = data[z][x][y]
            else:
                sampled_value = data[z][y][x]
            sampled_data.append((x, y, z, sampled_value))

    # 根据给定的步长和点的数量提取数据点
    if use_box_area:
        for step in step_sizes:
            x_range = use_box[1] // step
            y_range = use_box[2] // step
            z_range = use_box[3] // step
            for x in range(0, x_range + 1):
                for y in range(0, y_range + 1):
                    for z in range(0, z_range + 1):
                        x_coord = use_box[0][0]+x * step
                        y_coord = use_box[0][1]+y * step
                        z_coord = use_box[0][2]+z * step
                        # 检查是否满足离源距离条件
                        skip = False
                        for pos in sourcepos:
                            distance = np.sqrt((x_coord - pos[0])**2 + (y_coord - pos[1])**2 + (z_coord - pos[2])**2)
                            if distance <= 30:
                                skip = True
                                break
                        if not skip:
                            if x_y_reverse == False:
                                value = data[z_coord][x_coord][y_coord]
                            else:
                                value = data[z_coord][y_coord][x_coord]
                            sampled_data.append((x_coord, y_coord, z_coord, value))  
    else:    
        for step in step_sizes:
            x_range = inner_radius_max // step
            y_range = inner_radius_max // step
            z_range = inner_radius_max // step
            if direction == "6vector":
                for x in range(-x_range + 1, x_range + 1):
                    for y in range(-y_range + 1, y_range + 1):
                        for z in range(-z_range + 1, z_range + 1):
                            x_coord = center_x+x * step
                            y_coord = center_y+y * step
                            z_coord = center_z+z * step
                            
                            # 检查是否满足离源距离条件
                            skip = False
                            for pos in sourcepos:
                                distance = np.sqrt((x_coord - pos[0])**2 + (y_coord - pos[1])**2 + (z_coord - pos[2])**2)
                                if distance <= 10:
                                    skip = True
                                    break
                            if not skip:
                                if x_y_reverse == False:
                                    value = data[z_coord][x_coord][y_coord]
                                else:
                                    value = data[z_coord][y_coord][x_coord]
                                sampled_data.append((x_coord, y_coord, z_coord, value))
            elif direction == "3vector":
                for x in range(0, x_range + 1):
                    for y in range(0, y_range + 1):
                        for z in range(0, z_range + 1):
                            x_coord = center_x+x * step
                            y_coord = center_y+y * step
                            z_coord = center_z+z * step
                            # 检查是否满足离源距离条件
                            skip = False
                            for pos in sourcepos:
                                distance = np.sqrt((x_coord - pos[0])**2 + (y_coord - pos[1])**2 + (z_coord - pos[2])**2)
                                if distance <= 10:
                                    skip = True
                                    break
                            if not skip:
                                if x_y_reverse == False:
                                    value = data[z_coord][x_coord][y_coord]
                                else:
                                    value = data[z_coord][y_coord][x_coord]
                                sampled_data.append((x_coord, y_coord, z_coord, value))        

    
    # 添加源点（弃用/no）
    # sampled_data.append((63,58,29,data[29][58][63]))
    
    # 改格式为dataFrame
    sampled_df = pd.DataFrame(sampled_data, columns=['x', 'y', 'z', 'target'])
    return sampled_df

def testing_sampling(data, 
                     center_x, 
                     center_y, 
                     center_z, 
                     inner_radius, 
                     test_nums=90000, 
                     use_box_area = False, #是否使用固定空间来进行取样
                     use_box = [[0,0,0],74,99,99], #左前下点，x方向延申长度，y方向延申长度，z方向延申长度
                     x_y_reverse = True,
                     direction = "6vector",
                     sample_like_training = False,
                     step_sizes = [2,3,4]):
    if sample_like_training == True:
        return training_sampling(data, 
                      center_x, 
                      center_y, 
                      center_z, 
                      inner_radius, 
                      vertice = False, 
                      use_box_area = use_box_area, #是否使用固定空间来进行取样
                      use_box = use_box, #左前下点，x方向延申长度，y方向延申长度，z方向延申长度
                      inner_radius_list = list(range(1,3)), 
                      step_sizes = step_sizes, 
                      x_y_reverse = x_y_reverse, 
                      direction = direction)
    if direction == "6vector":
        vertices = [(random.randint(center_x-inner_radius, center_x+inner_radius), random.randint(center_y-inner_radius, center_y+inner_radius), random.randint(center_z-inner_radius, center_z+inner_radius)) for i in range(test_nums)]
    elif direction == "3vector":
        vertices = [(random.randint(center_x, center_x+inner_radius), random.randint(center_y, center_y+inner_radius), random.randint(center_z, center_z+inner_radius)) for i in range(test_nums)]
    sampled_data = []
    for vertex in vertices:
        x, y, z = vertex
        if x_y_reverse == False: 
            sampled_value = data[z][x][y]
        else:
            sampled_value = data[z][y][x]
            
        sampled_data.append((x, y, z, sampled_value))
        
    sampled_df = pd.DataFrame(sampled_data, columns=['x', 'y', 'z', 'target'])
    return sampled_df


    
    
    
# r'X coordinate: from.*?(\d+) bins.*?\nY coordinate: from.*?(\d+) bins.*?\nZ coordinate: from.*?(\d+) bins'